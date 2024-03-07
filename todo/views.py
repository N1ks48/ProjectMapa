from django.shortcuts import render
import cx_Oracle
from django.http import JsonResponse, HttpResponse
from openpyxl import Workbook
import ast
from datetime import datetime


def execute_sql(sql, **kwargs):
    try:
        # Установка соединения с Oracle
        connection = cx_Oracle.connect("NSAR", "L:FYUJAHB", "192.168.201.1:1521/Primus", encoding="UTF-8")
        # Создание курсора
        cursor = connection.cursor()
        # Выполнение SQL-запроса
        cursor.execute(sql, kwargs)
        # Получение результатов
        results = cursor.fetchall()
        # Фиксация изменений и закрытие соединения
        cursor.close()
        connection.close()
        # Возврат значения
        return results

    except Exception as e:
        # Обработка ошибок
        return f"Error: {str(e)}"


def get_option():
    try:
        qwerty = '''select okpo, name from NAU_TEST.FIRMS where comments = 'ю' '''

        # Получение результатов
        options = execute_sql(qwerty)
        # Фиксация изменений и закрытие соединения

        html_options = ""
        for key, name in options:
            html_options += f'<option value="{name}">{key}</option>'
        return html_options
    except cx_Oracle.Error as error:
        # Обработка ошибок подключения к базе данных
        print(f"Error connecting to Oracle Database: {error}")


def get_filters(request):
    filters = {}
    switches = []
    if request.POST:
        aptid = request.POST.get('aptid')
        name = request.POST.get('name')
        city = request.POST.get('city')
        okpo = request.POST.get('okpo')
        switch1 = request.POST.get('Switch1')
        switch2 = request.POST.get('Switch2')
        switch3 = request.POST.get('Switch3')
        switch4 = request.POST.get('Switch4')

        if aptid:
            filters['f.ID'] = aptid
        if name:
            filters['f.NAME'] = name
        if city:
            filters['f.CITY'] = city
        if okpo:
            filters['ff.NAME'] = okpo

        if switch1:
            switches.append(2)
        if switch2:
            switches.append(3)
        if switch3:
            switches.append(4)
        if switch4:
            switches.append(7)

    if switches:
        filters['st.STATUSAPTEKAID'] = switches

    return filters


def build_conditions(filters):
    conditions = []
    for key, value in filters.items():
        if value:
            if isinstance(value, list):  # Для обработки switches
                conditions.append(f"{key} IN ({', '.join(map(str, value))})")
            else:
                conditions.append(f"upper({key}) like upper('%{value}%')")
    return conditions


def map_view(request):
    try:
        sql = '''SELECT f.NAME, f.MAP_LAT, f.MAP_LNG, sa.NAME, f.CITY, ff.NAME FROM NAU_TEST.FIRMS f
                    LEFT JOIN nau_test.firms ff ON ff.id = f.PARENTID
                    LEFT JOIN (SELECT f.id firmid, CASE WHEN tc.isclosed = 1 THEN 7 ELSE f.statusaptekaid END statusaptekaid
                               FROM NAU_TEST.FIRMS f 
                               LEFT JOIN nau_test.firms_temprory_closed tc ON tc.firmsid = f.id) st ON st.firmid = f.id
                    LEFT JOIN nau_test.STATUSAPTEKA sa ON sa.id = st.STATUSAPTEKAID
                    WHERE f.MAP_LAT IS NOT NULL AND f.city IS NOT NULL AND st.STATUSAPTEKAID IN (2, 3, 4, 7)
                    AND f.id != 501413 and trim(f.CITY) like trim('%Київ%') '''

        rows = execute_sql(sql)
        html_options = get_option()

        locations = []
        for result in rows:
            name, latitude, longitude, status, city, okpo = result
            locations.append({"name": name, "latitude": latitude, "longitude": longitude, "status": status,
                              "city": city, "okpo": okpo})
        context = {'html_options': html_options, 'locations': locations}
    except Exception as e:
        print(f"Error: {e}")
        context = {'html_options': [], 'locations': []}

    return render(request, 'mymap/map.html', context)


def maps_filter(request):
    if request.method == 'POST':
        filters = get_filters(request)
        conditions = build_conditions(filters)

        qwr_base = '''SELECT f.NAME, f.MAP_LAT, f.MAP_LNG, sa.NAME, f.CITY, ff.NAME FROM NAU_TEST.FIRMS f
                    LEFT JOIN nau_test.firms ff ON ff.id = f.PARENTID
                    LEFT JOIN (SELECT f.id firmid, CASE WHEN tc.isclosed = 1 THEN 7 
                    ELSE f.statusaptekaid END statusaptekaid
                    FROM NAU_TEST.FIRMS f 
                    LEFT JOIN nau_test.firms_temprory_closed tc ON tc.firmsid = f.id) st ON st.firmid = f.id
                    LEFT JOIN nau_test.STATUSAPTEKA sa ON sa.id = st.STATUSAPTEKAID
                    WHERE f.MAP_LAT IS NOT NULL AND f.city IS NOT NULL AND st.STATUSAPTEKAID IN (2, 3, 4, 7)
                    AND f.id != 501413'''

        if conditions:
            qwr_base += f" AND {' AND '.join(conditions)}"
        print(qwr_base)
        try:
            results = execute_sql(qwr_base)
            html_options = get_option()
            locations = []
            for result in results:
                name, latitude, longitude, status, city, okpo = result
                locations.append({"name": name, "latitude": latitude, "longitude": longitude, "status": status,
                                  "city": city, "okpo": okpo})
            context = {'html_options': html_options, 'locations': locations}
            return render(request, 'mymap/map.html', context)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)
    else:
        return render(request, 'mymap/map.html')


def changes_map(request):
    if request.method == 'POST':
        filters = get_filters(request)
        conditions = build_conditions(filters)

        if not filters:
            sql = '''SELECT f.id, f.name, f.okpo, f.address2, f.city, sa.NAME FROM NAU_TEST.FIRMS f
                        LEFT JOIN nau_test.firms ff ON ff.id = f.PARENTID
                        LEFT JOIN (SELECT f.id firmid, CASE WHEN tc.isclosed = 1 THEN 7 
                        ELSE f.statusaptekaid END statusaptekaid
                        FROM NAU_TEST.FIRMS f 
                        LEFT JOIN nau_test.firms_temprory_closed tc ON tc.firmsid = f.id) st ON st.firmid = f.id
                        LEFT JOIN nau_test.STATUSAPTEKA sa ON sa.id = st.STATUSAPTEKAID
                        WHERE st.STATUSAPTEKAID IN (2, 3, 4, 7)'''
        else:
            conditions_str = ' AND '.join(conditions)
            sql = f'''SELECT f.id, f.name, f.okpo, f.address2, f.city, sa.NAME FROM NAU_TEST.FIRMS f
                        LEFT JOIN nau_test.firms ff ON ff.id = f.PARENTID
                        LEFT JOIN (SELECT f.id firmid, CASE WHEN tc.isclosed = 1 THEN 7 
                        ELSE f.statusaptekaid END statusaptekaid
                        FROM NAU_TEST.FIRMS f 
                        LEFT JOIN nau_test.firms_temprory_closed tc ON tc.firmsid = f.id) st ON st.firmid = f.id
                        LEFT JOIN nau_test.STATUSAPTEKA sa ON sa.id = st.STATUSAPTEKAID 
                        WHERE st.STATUSAPTEKAID IN (2, 3, 4, 7) AND {conditions_str}'''
        print(sql)
        try:
            results = execute_sql(sql)
            html_options = get_option()
            context = {'html_options': html_options, 'results': results}
            return render(request, 'mymap/changes_map.html', context)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)
    else:
        html_options = get_option()
        return render(request, 'mymap/changes_map.html', {'html_options': html_options})


def download_data(request):
    if request.method == 'POST':
        try:
            # Получение строки результатов из запроса
            raw_results_string = request.POST.getlist('results')

            # Разбивка строки на строки (по символу новой строки)
            rows = [row.strip() for row in raw_results_string]

            # Создание нового XLSX-файла
            workbook = Workbook()
            worksheet = workbook.active

            # Заголовки колонок
            headers = ['ID', 'Назва', 'ЕРДПОУ', 'Адреса', 'Місто']

            # Запись заголовков в первую строку
            for col_num, header in enumerate(headers, start=1):
                worksheet.cell(row=1, column=col_num, value=header)

            # Запись данных в соответствующие колонки
            for row_num, row_data in enumerate(rows, start=2):
                # Используем ast.literal_eval для безопасного преобразования строки в кортеж
                row_tuple = ast.literal_eval(row_data)
                # Запись данных в ячейки с учетом разных типов данных
                for col_num, value in enumerate(row_tuple, start=1):
                    worksheet.cell(row=row_num, column=col_num, value=value)

            current_datetime = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')

            # Создание HTTP-ответа с XLSX-файлом
            filename = f"filtered_data_{current_datetime}.xlsx"
            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            workbook.save(response)

            return response
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=400)
